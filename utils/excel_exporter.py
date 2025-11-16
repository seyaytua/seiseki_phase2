import pandas as pd
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelExporter:
    """Excel出力クラス"""
    
    def __init__(self):
        self.default_export_dir = Path('data/exports')
        self.default_export_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_excel(self, data, columns, filename, sheet_name='Sheet1'):
        """データをExcelファイルに出力"""
        try:
            df = pd.DataFrame(data)
            
            available_columns = [col for col in columns if col in df.columns]
            if available_columns:
                df = df[available_columns]
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = Path(filename).stem
            export_filename = f"{base_name}_{timestamp}.xlsx"
            export_path = self.default_export_dir / export_filename
            
            with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # ヘッダーのスタイル
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 罫線
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                               min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border
                
                # 列幅の自動調整
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            return str(export_path)
        
        except Exception as e:
            raise Exception(f"Excel出力エラー: {str(e)}")
    
    def export_multiple_sheets(self, data_dict, filename):
        """複数のシートを持つExcelファイルを出力"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = Path(filename).stem
            export_filename = f"{base_name}_{timestamp}.xlsx"
            export_path = self.default_export_dir / export_filename
            
            with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
                for sheet_name, (data, columns) in data_dict.items():
                    df = pd.DataFrame(data)
                    
                    available_columns = [col for col in columns if col in df.columns]
                    if available_columns:
                        df = df[available_columns]
                    
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    worksheet = writer.sheets[sheet_name]
                    
                    # ヘッダーのスタイル
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 罫線
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                                   min_col=1, max_col=worksheet.max_column):
                        for cell in row:
                            cell.border = thin_border
                    
                    # 列幅の自動調整
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            return str(export_path)
        
        except Exception as e:
            raise Exception(f"Excel出力エラー: {str(e)}")