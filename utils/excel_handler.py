import pandas as pd
from openpyxl import load_workbook


class ExcelHandler:
    @staticmethod
    def read_excel_preview(file_path, header_row=0, max_rows=10):
        """Excelファイルのプレビューを読み込む"""
        try:
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            df = pd.read_excel(
                file_path,
                sheet_name=0,
                header=header_row,
                nrows=max_rows
            )
            
            return {
                'success': True,
                'sheet_names': sheet_names,
                'preview_data': df,
                'columns': list(df.columns)
            }
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    @staticmethod
    def read_single_sheet(file_path, sheet_name, header_row=0):
        """単一シートを読み込む"""
        try:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header_row
            )
            return df
        except Exception as e:
            raise Exception(f"シート読み込みエラー: {str(e)}")
    
    @staticmethod
    def read_multiple_sheets(file_path, header_row=0, exclude_sheets=None):
        """複数シートを読み込んで統合"""
        try:
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            if exclude_sheets:
                sheet_names = [s for s in sheet_names if s not in exclude_sheets]
            
            dfs = []
            for sheet_name in sheet_names:
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    header=header_row
                )
                df['_source_sheet'] = sheet_name
                dfs.append(df)
            
            combined_df = pd.concat(dfs, ignore_index=True)
            return combined_df
        except Exception as e:
            raise Exception(f"複数シート読み込みエラー: {str(e)}")
    
    @staticmethod
    def get_sheet_names(file_path):
        """シート名のリストを取得"""
        try:
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            raise Exception(f"シート名取得エラー: {str(e)}")