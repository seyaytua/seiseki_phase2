import pandas as pd
from pathlib import Path
from openpyxl import load_workbook


class MultiSheetHandler:
    """複数シート統合処理（最適化版）"""
    
    @staticmethod
    def convert_sheets_to_vertical(file_path, header_row=0, exclude_sheets=None):
        """
        複数シートを縦データに統合
        
        Args:
            file_path: Excelファイルパス
            header_row: ヘッダー行番号
            exclude_sheets: 除外するシート名のリスト
        
        Returns:
            DataFrame: 統合されたデータ
        """
        try:
            print(f"ファイル読み込み中: {Path(file_path).name}")
            
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            print(f"シート数: {len(sheet_names)}")
            
            if exclude_sheets:
                sheet_names = [s for s in sheet_names if s not in exclude_sheets]
                print(f"除外後のシート数: {len(sheet_names)}")
            
            dfs = []
            total_rows = 0
            
            for idx, sheet_name in enumerate(sheet_names):
                print(f"シート {idx+1}/{len(sheet_names)} 読み込み中: {sheet_name}")
                
                try:
                    df = pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        header=header_row
                    )
                    
                    # 空のDataFrameはスキップ
                    if df.empty:
                        print(f"  空のシート - スキップ")
                        continue
                    
                    print(f"  読み込み完了: {len(df)}行")
                    
                    # シート名をカラムとして追加
                    df['_source_sheet'] = sheet_name
                    dfs.append(df)
                    total_rows += len(df)
                    
                    print(f"  累計: {total_rows}行")
                    
                except Exception as e:
                    print(f"  シート読み込みエラー: {e}")
                    continue
            
            if not dfs:
                raise Exception("有効なデータを含むシートが見つかりませんでした")
            
            print(f"シート統合中...")
            # 全シートを縦に結合
            combined_df = pd.concat(dfs, ignore_index=True)
            print(f"統合完了: 総行数 {len(combined_df)}行")
            
            return combined_df
        
        except Exception as e:
            raise Exception(f"複数シート統合エラー: {str(e)}")
    
    @staticmethod
    def merge_multiple_files(file_paths, header_row=0, output_path=None):
        """
        複数ファイルを1つに統合（最適化版）
        
        Args:
            file_paths: ファイルパスのリスト
            header_row: ヘッダー行番号
            output_path: 出力先パス（省略時は自動生成）
        
        Returns:
            str: 出力されたファイルパス
        """
        try:
            print(f"\n=== 複数ファイル統合開始 ===")
            print(f"ファイル数: {len(file_paths)}")
            
            # 中間ファイルリスト
            temp_files = []
            temp_dir = Path("data/temp")
            temp_dir.mkdir(parents=True, exist_ok=True)
            
            for idx, file_path in enumerate(file_paths):
                print(f"\n[{idx+1}/{len(file_paths)}] 処理中: {Path(file_path).name}")
                
                try:
                    # 各ファイルをシート統合
                    df = MultiSheetHandler.convert_sheets_to_vertical(
                        file_path,
                        header_row
                    )
                    
                    # ファイル名をカラムとして追加
                    df['_source_file'] = Path(file_path).name
                    
                    # 中間ファイルとして保存（メモリ節約）
                    temp_path = temp_dir / f"temp_{idx}.xlsx"
                    print(f"中間ファイル保存中...")
                    df.to_excel(temp_path, index=False, engine='openpyxl')
                    temp_files.append(temp_path)
                    
                    print(f"✓ ファイル処理完了: {len(df)}行")
                    
                    # メモリ解放
                    del df
                    
                except Exception as e:
                    print(f"❌ ファイル処理エラー: {e}")
                    continue
            
            if not temp_files:
                raise Exception("有効なデータを含むファイルが見つかりませんでした")
            
            print(f"\n中間ファイル統合中...")
            # 中間ファイルを読み込んで統合
            all_dfs = []
            for idx, temp_file in enumerate(temp_files):
                print(f"中間ファイル {idx+1}/{len(temp_files)} 読み込み中...")
                df = pd.read_excel(temp_file)
                all_dfs.append(df)
            
            print(f"全データ統合中...")
            final_df = pd.concat(all_dfs, ignore_index=True)
            print(f"統合完了: 総行数 {len(final_df)}行")
            
            # 出力先パス生成
            if output_path is None:
                from datetime import datetime
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"data/merged/merged_{timestamp}.xlsx"
            
            # 出力先ディレクトリ作成
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            
            print(f"Excelファイル出力中: {Path(output_path).name}")
            final_df.to_excel(output_path, index=False, engine='openpyxl')
            print(f"✓ 出力完了")
            
            # 中間ファイル削除
            print(f"中間ファイル削除中...")
            for temp_file in temp_files:
                try:
                    temp_file.unlink()
                except Exception as e:
                    print(f"中間ファイル削除エラー: {e}")
            
            # 一時ディレクトリ削除
            try:
                temp_dir.rmdir()
            except:
                pass
            
            return output_path
        
        except Exception as e:
            # エラー時も中間ファイルを削除
            try:
                temp_dir = Path("data/temp")
                if temp_dir.exists():
                    for temp_file in temp_dir.glob("temp_*.xlsx"):
                        temp_file.unlink()
                    temp_dir.rmdir()
            except:
                pass
            
            raise Exception(f"複数ファイル統合エラー: {str(e)}")
    
    @staticmethod
    def get_sheet_names(file_path):
        """シート名のリストを取得"""
        try:
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            raise Exception(f"シート名取得エラー: {str(e)}")